#!/usr/bin/env python3
"""
Extract all data from LAMP_List_2026.xlsx
"""

import openpyxl
import json
import sys

def extract_excel_data(filepath):
    """Extract all sheets from Excel file"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = []

        # Get headers from first row
        headers = None
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx == 1:
                headers = [cell if cell is not None else f"col_{i}" for i, cell in enumerate(row)]
                continue

            # Create row dict
            row_dict = {}
            for col_idx, (header, cell) in enumerate(zip(headers, row)):
                if cell is not None:
                    # Handle various data types
                    if isinstance(cell, str):
                        row_dict[header] = cell
                    elif isinstance(cell, (int, float, bool)):
                        row_dict[header] = cell
                    elif cell is not None:
                        row_dict[header] = str(cell)

            if row_dict:  # Only add non-empty rows
                sheet_data.append(row_dict)

        if sheet_data:
            data[sheet_name] = {
                'headers': headers,
                'rows': sheet_data,
                'count': len(sheet_data)
            }

    return data

if __name__ == '__main__':
    import os
    excel_file = os.environ.get('LAMP_EXCEL_FILE', './LAMP_List_2026.xlsx')

    print(f"📂 Reading: {excel_file}\n")
    data = extract_excel_data(excel_file)

    print(f"📊 Found {len(data)} sheet(s) with data:\n")
    for sheet_name, sheet_data in data.items():
        print(f"   {sheet_name}: {sheet_data['count']} rows")
        if sheet_data['headers']:
            print(f"      Columns: {', '.join([str(h) for h in sheet_data['headers'][:10]])}")
            if len(sheet_data['headers']) > 10:
                print(f"                ... and {len(sheet_data['headers']) - 10} more")
        print()

    # Save to JSON
    output_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'exports', 'lamp-excel-data.json')
    with open(output_file, 'w') as f:
        json.dump(data, f, indent=2, default=str)

    print(f"✅ Data saved to: {output_file}")
    print(f"📏 File size: {len(json.dumps(data)) / 1024:.2f} KB")
